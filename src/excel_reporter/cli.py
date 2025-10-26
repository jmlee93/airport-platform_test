"""명령행 인터페이스와 기본 분석 로직을 제공합니다."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook


def analyze_workbook(path: str | Path) -> Dict[str, Any]:
    """주어진 엑셀 파일을 열어 시트와 수식 요약 정보를 반환합니다.

    현재는 MVP 수준으로, 다음 정보를 제공합니다.
    * 워크북의 모든 시트 이름
    * 각 시트에 포함된 수식 셀 수
    * 첫 번째 5개의 수식 예시

    추후에 수식 그래프 구성, 민감도 분석 등으로 확장할 수 있습니다.
    """

    workbook = load_workbook(filename=path, data_only=False)
    summary: Dict[str, Any] = {"workbook": Path(path).name, "sheets": []}

    for sheet in workbook.worksheets:
        formulas = [cell.coordinate for row in sheet.iter_rows() for cell in row if cell.data_type == "f"]
        sheet_info = {
            "name": sheet.title,
            "formula_count": len(formulas),
            "formula_samples": formulas[:5],
        }
        summary["sheets"].append(sheet_info)

    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="엑셀 재무 모델 요약 리포트 생성")
    parser.add_argument("excel_file", type=Path, help="분석할 엑셀 파일 경로")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        help="요약 정보를 저장할 JSON 파일 경로 (미지정 시 표준 출력)",
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)

    summary = analyze_workbook(args.excel_file)

    if args.output:
        args.output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
