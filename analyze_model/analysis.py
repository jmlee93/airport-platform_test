"""Core spreadsheet analysis routines."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import AnalysisConfig
from .logging_utils import get_error_logger


@dataclass
class SheetMetrics:
    """Summary information about a single worksheet."""

    name: str
    total_cells: int
    formula_cells: int
    empty_cells: int
    error_cells: int
    issues: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, object]:
        return {
            "name": self.name,
            "total_cells": self.total_cells,
            "formula_cells": self.formula_cells,
            "empty_cells": self.empty_cells,
            "error_cells": self.error_cells,
            "issues": list(self.issues),
        }


@dataclass
class AnalysisResult:
    """Aggregate outcome of an analysis run."""

    workbook: Path
    generated_at: datetime
    sheet_metrics: Dict[str, SheetMetrics]
    missing_sheets: List[str]
    parse_errors: List[str]
    metadata: Dict[str, str]

    def to_dict(self) -> Dict[str, object]:
        return {
            "workbook": str(self.workbook),
            "generated_at": self.generated_at.isoformat(),
            "sheets": {name: metrics.to_dict() for name, metrics in self.sheet_metrics.items()},
            "missing_sheets": list(self.missing_sheets),
            "parse_errors": list(self.parse_errors),
            "metadata": dict(self.metadata),
        }


def _iter_cells(sheet: Worksheet):
    for row in sheet.iter_rows():
        for cell in row:
            yield cell


def _analyze_sheet(sheet: Worksheet) -> SheetMetrics:
    total_cells = sheet.max_row * sheet.max_column
    formula_cells = 0
    empty_cells = 0
    error_cells = 0
    issues: List[str] = []

    for cell in _iter_cells(sheet):
        value = cell.value
        if value is None or value == "":
            empty_cells += 1
        if cell.data_type == "f":
            formula_cells += 1
        if cell.data_type == "e":
            error_cells += 1
            issues.append(
                f"Cell {cell.coordinate} contains an Excel error ({value})."
            )
        if isinstance(value, str) and value.startswith("=#REF!"):
            issues.append(
                f"Cell {cell.coordinate} contains an unresolved reference ({value})."
            )

    if total_cells == 0:
        issues.append("Worksheet is empty.")

    return SheetMetrics(
        name=sheet.title,
        total_cells=total_cells,
        formula_cells=formula_cells,
        empty_cells=empty_cells,
        error_cells=error_cells,
        issues=issues,
    )


def analyze_workbook(
    workbook_path: Path,
    config: AnalysisConfig,
    audit_logger=None,
) -> AnalysisResult:
    """Analyze an Excel workbook and produce a structured report."""

    workbook_path = Path(workbook_path)
    sheet_metrics: Dict[str, SheetMetrics] = {}
    missing_sheets: List[str] = []
    parse_errors: List[str] = []

    if audit_logger is not None:
        audit_logger.info("Starting analysis for %s", workbook_path)

    try:
        workbook = load_workbook(workbook_path, data_only=False)
    except Exception as exc:  # pragma: no cover - defensive logging
        message = f"Failed to open workbook: {exc}"
        parse_errors.append(message)
        if audit_logger is not None:
            audit_logger.error(message)
        get_error_logger().exception("Workbook parsing failure for %s", workbook_path)
        raise

    seen_sheets = set()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        seen_sheets.add(sheet_name)
        metrics = _analyze_sheet(sheet)
        sheet_metrics[sheet_name] = metrics
        if metrics.issues and audit_logger is not None:
            for issue in metrics.issues:
                audit_logger.warning("%s | %s", sheet_name, issue)

    expected = set(config.expected_sheets)
    for expected_sheet in expected:
        if expected_sheet not in seen_sheets:
            missing_sheets.append(expected_sheet)
            if audit_logger is not None:
                audit_logger.error("Missing expected sheet: %s", expected_sheet)

    result = AnalysisResult(
        workbook=workbook_path,
        generated_at=datetime.utcnow(),
        sheet_metrics=sheet_metrics,
        missing_sheets=missing_sheets,
        parse_errors=parse_errors,
        metadata=config.metadata,
    )

    if audit_logger is not None:
        audit_logger.info(
            "Completed analysis for %s with %d sheets and %d issues",
            workbook_path,
            len(sheet_metrics),
            sum(len(metrics.issues) for metrics in sheet_metrics.values()) + len(missing_sheets),
        )

    return result


__all__ = [
    "AnalysisResult",
    "SheetMetrics",
    "analyze_workbook",
]
