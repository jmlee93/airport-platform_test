# analysis/cli.py
# Minimal CLI: scan workbook for candidate input cells
import argparse, json, csv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill

def _is_yellow(cell) -> bool:
    try:
        fill: PatternFill | None = cell.fill
        if not fill or not fill.fgColor:
            return False
        c: Color = fill.fgColor
        rgb = (c.rgb or "").upper()       # AARRGGBB 또는 RRGGBB
        return rgb.endswith("FFFF00") or rgb in {"FFFF00", "FFFFFF00", "FFFBF2CC"}
    except Exception:
        return False

def _looks_like_input(cell) -> bool:
    txt = (str(cell.value) if cell.value is not None else "").strip()
    cmt = (cell.comment.text if cell.comment else "").lower()
    return (
        txt.startswith("INPUT_")
        or "sensitivity" in cmt
        or "input" in cmt
        or _is_yellow(cell)
    )

def scan(wb_path: Path):
    wb = load_workbook(filename=str(wb_path), data_only=False, keep_links=True)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _looks_like_input(cell):
                    label = None
                    if cell.col_idx > 1:
                        left = ws.cell(row=cell.row, column=cell.col_idx - 1).value
                        if isinstance(left, str):
                            label = left.strip()
                    out.append({
                        "sheet": ws.title,
                        "addr": cell.coordinate,
                        "label": label,
                        "value": cell.value,
                        "comment": cell.comment.text if cell.comment else None
                    })
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--wb", required=True, help="path to .xlsx/.xlsm")
    ap.add_argument("--out", required=True, help="output JSON or CSV path")
    args = ap.parse_args()

    items = scan(Path(args.wb))
    outp = Path(args.out)
    outp.parent.mkdir(parents=True, exist_ok=True)

    if outp.suffix.lower() == ".csv":
        with outp.open("w", newline="", encoding="utf-8") as f:
            fields = ["sheet","addr","label","value","comment"]
            w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(items)
    else:
        outp.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {len(items)} items -> {outp}")

if __name__ == "__main__":
    main()
