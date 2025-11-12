"""Utilities for discovering sensitivity input cells within Excel workbooks.

This module exposes :func:`extract_sensitivity_inputs`, which traverses an Excel
workbook and returns the cells that have been tagged for sensitivity analysis
via naming conventions, cell highlighting, or cell comments.  The helper is
primarily intended for analytical pipelines where the set of adjustable inputs
must be enumerated before executing simulations.
"""
from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import List, Optional, Sequence, Union

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook

try:  # Optional dependency – only required when ``output_format="dataframe"``
    import pandas as _pd  # type: ignore
except Exception:  # pragma: no cover - pandas is optional for library users
    _pd = None


ColorLike = Union[str, "openpyxl.styles.colors.Color", None]


@dataclass(frozen=True)
class SensitivityCell:
    """Structured metadata describing a sensitivity input cell."""

    sheet: str
    coordinate: str
    value: Optional[object]
    parameter: Optional[str]
    comment: Optional[str]
    fill_color: Optional[str]
    matched_by: Sequence[str]

    def as_dict(self) -> dict:
        return {
            "sheet": self.sheet,
            "coordinate": self.coordinate,
            "value": self.value,
            "parameter": self.parameter,
            "comment": self.comment,
            "fill_color": self.fill_color,
            "matched_by": list(self.matched_by),
        }


def extract_sensitivity_inputs(
    workbook: Union[str, Path, Workbook],
    *,
    prefix: str = "Sensitivity_",
    comment_keywords: Optional[Sequence[str]] = None,
    highlight_colors: Optional[Sequence[ColorLike]] = None,
    output_format: str = "records",
    data_only: bool = False,
) -> Union[List[dict], str, "_pd.DataFrame"]:
    """Identify the cells tagged as sensitivity inputs within an Excel workbook.

    Parameters
    ----------
    workbook:
        Either a path to an Excel workbook on disk or an :class:`openpyxl`
        :class:`~openpyxl.workbook.workbook.Workbook` instance.
    prefix:
        String prefix applied to sensitivity input names (defaults to
        ``"Sensitivity_"``).  When a cell's value begins with this prefix the
        remainder of the value is treated as the parameter name.
    comment_keywords:
        Keywords that, if present within a cell's comment text, qualify the cell
        as a sensitivity input.  The ``prefix`` is used when this argument is
        omitted.
    highlight_colors:
        Sequence of colors that qualify a cell as a sensitivity input when its
        fill color matches.  Colors can be provided as 6- or 8-digit hex strings
        (with or without the leading ``#``) or as
        :class:`openpyxl.styles.colors.Color` instances.
    output_format:
        One of ``"records"`` (default), ``"json"``, or ``"dataframe"``.
    data_only:
        Passed through to :func:`openpyxl.load_workbook` when ``workbook`` is a
        path.  Defaults to ``False`` so that formulas are not evaluated.

    Returns
    -------
    list of dict | str | pandas.DataFrame
        Depending on ``output_format``, returns a list of dictionaries (records
        format), a JSON string, or a :class:`pandas.DataFrame`.
    """

    if isinstance(workbook, (str, Path)):
        wb = load_workbook(filename=str(workbook), data_only=data_only)
    elif isinstance(workbook, Workbook):
        wb = workbook
    else:  # pragma: no cover - defensive branch
        raise TypeError(
            "workbook must be a path or an openpyxl Workbook, got "
            f"{type(workbook)!r}"
        )

    keywords = tuple(comment_keywords or [prefix])
    normalized_highlights = (
        tuple(_normalize_color(color) for color in highlight_colors)
        if highlight_colors
        else tuple()
    )

    sensitivity_cells: List[SensitivityCell] = []
    for worksheet in wb.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                reasons = _matching_reasons(
                    cell,
                    prefix=prefix,
                    keywords=keywords,
                    highlight_colors=normalized_highlights,
                )
                if not reasons:
                    continue

                parameter = _derive_parameter_name(cell, prefix)
                sensitivity_cells.append(
                    SensitivityCell(
                        sheet=worksheet.title,
                        coordinate=cell.coordinate,
                        value=cell.value,
                        parameter=parameter,
                        comment=cell.comment.text if cell.comment else None,
                        fill_color=_normalize_color(cell.fill.fgColor),
                        matched_by=tuple(sorted(reasons)),
                    )
                )

    records = [cell.as_dict() for cell in sensitivity_cells]

    if output_format == "records":
        return records
    if output_format == "json":
        return json.dumps(records, ensure_ascii=False, indent=2)
    if output_format == "dataframe":
        if _pd is None:
            raise ImportError(
                "pandas is required when output_format='dataframe' but could not"
                " be imported."
            )
        return _pd.DataFrame.from_records(records)

    raise ValueError(
        "output_format must be one of {'records', 'json', 'dataframe'}"
    )


def _matching_reasons(
    cell: Cell,
    *,
    prefix: str,
    keywords: Sequence[str],
    highlight_colors: Sequence[Optional[str]],
) -> List[str]:
    reasons: List[str] = []

    if isinstance(cell.value, str) and cell.value.startswith(prefix):
        reasons.append("prefix")

    if cell.comment is not None:
        comment_text = cell.comment.text or ""
        if any(keyword in comment_text for keyword in keywords):
            reasons.append("comment")

    cell_color = _normalize_color(cell.fill.fgColor)
    if highlight_colors and cell_color in highlight_colors:
        reasons.append("fill_color")

    return reasons


def _derive_parameter_name(cell: Cell, prefix: str) -> Optional[str]:
    if isinstance(cell.value, str) and cell.value.startswith(prefix):
        raw = cell.value[len(prefix) :]
        return _clean_parameter(raw)
    if cell.comment:
        text = cell.comment.text or ""
        for line in text.splitlines():
            if line.startswith(prefix):
                return _clean_parameter(line[len(prefix) :])
    return None


def _normalize_color(color: ColorLike) -> Optional[str]:
    """Normalize openpyxl color values to a consistent hex string."""

    if color is None:
        return None

    rgb = None
    try:
        rgb = getattr(color, "rgb", None)
    except Exception:  # pragma: no cover - defensive: getattr failure
        rgb = None

    if rgb:
        value = rgb
    else:
        value = str(color)

    if value is None:
        return None

    value = value.strip().lstrip("#").upper()

    if len(value) == 8 and value.startswith("FF"):
        value = value[2:]

    if not value:
        return None

    return value


def _clean_parameter(text: str) -> Optional[str]:
    candidate = text.strip()
    if not candidate:
        return None
    if ":" in candidate:
        head, *_ = candidate.split(":", 1)
        head = head.strip()
        if head:
            return head
    return candidate


__all__ = ["extract_sensitivity_inputs", "SensitivityCell"]
