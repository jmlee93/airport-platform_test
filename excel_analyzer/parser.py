from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from numbers import Number
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import get_column_letter

from .models import CellMetadata, NamedRangeMetadata, SheetMetadata, TableMetadata, WorkbookMetadata


def normalize_coordinate(coordinate: str) -> str:
    """Remove absolute markers from a coordinate."""

    return coordinate.replace("$", "")


def parse_reference(ref: str, current_sheet: str) -> List[str]:
    """Expand a cell or range reference into workbook-qualified cell ids."""

    if "!" in ref:
        sheet_name, coord = ref.split("!", 1)
        sheet_name = sheet_name.strip("'")
    else:
        sheet_name = current_sheet
        coord = ref

    coord = coord.replace("$", "")

    if ":" not in coord:
        return [f"{sheet_name}!{coord}"]

    min_col, min_row, max_col, max_row = range_boundaries(coord)
    cells: List[str] = []
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        for row in range(min_row, max_row + 1):
            cells.append(f"{sheet_name}!{col_letter}{row}")
    return cells


def extract_dependencies(formula: str, current_sheet: str) -> List[str]:
    """Parse a formula and extract all referenced cell ids."""

    tokenizer = Tokenizer(formula)
    dependencies: List[str] = []
    for token in tokenizer.items:
        if token.subtype in {"RANGE", "REF"}:
            try:
                dependencies.extend(parse_reference(token.value, current_sheet))
            except ValueError:
                # Ignore references we cannot parse (structured references, external workbooks)
                continue
    return sorted(set(dependencies))


def collect_named_ranges(workbook) -> List[NamedRangeMetadata]:
    named_ranges: List[NamedRangeMetadata] = []
    for defined_name in workbook.defined_names.definedName:
        destinations = []
        if defined_name.type == "RANGE":
            for sheet_name, target in defined_name.destinations:
                destinations.append(f"{sheet_name}!{normalize_coordinate(target)}")
        named_ranges.append(NamedRangeMetadata(name=defined_name.name, targets=destinations))
    return named_ranges


def dataframe_from_range(sheet, ref: str) -> pd.DataFrame:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    data: List[List[Optional[object]]] = []
    for row in sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=True):
        data.append(list(row))
    if not data:
        return pd.DataFrame()
    header, *rows = data
    return pd.DataFrame(rows, columns=header)


def range_has_merged_cells(sheet, ref: str) -> bool:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    for merged_range in sheet.merged_cells.ranges:
        m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged_range))
        if not (m_max_col < min_col or m_min_col > max_col or m_max_row < min_row or m_min_row > max_row):
            return True
    return False


def detect_structured_ranges(sheet, value_sheet) -> List[TableMetadata]:
    structured_ranges: List[TableMetadata] = []
    values = list(value_sheet.values)
    num_rows = len(values)
    num_cols = max((len(row) for row in values), default=0)

    def row_values(idx: int) -> List[Optional[object]]:
        if idx < 0 or idx >= num_rows:
            return []
        row = values[idx]
        return list(row) + [None] * (num_cols - len(row))

    row = 0
    while row < num_rows:
        current_row = row_values(row)
        non_empty_cells = [cell for cell in current_row if cell not in (None, "")]
        if len(non_empty_cells) >= 2 and all(isinstance(cell, str) for cell in non_empty_cells):
            next_row = row_values(row + 1)
            if any(
                isinstance(cell, Number) or isinstance(cell, (date, datetime))
                for cell in next_row
                if cell not in (None, "")
            ):
                start_col = next(i for i, cell in enumerate(current_row) if cell not in (None, ""))
                end_col = len(current_row) - 1
                while end_col > start_col and current_row[end_col] in (None, ""):
                    end_col -= 1
                end_row = row + 1
                last_data_row = end_row
                while end_row < num_rows:
                    values_row = row_values(end_row)
                    if all(values_row[col] in (None, "") for col in range(start_col, end_col + 1)):
                        break
                    last_data_row = end_row
                    end_row += 1
                ref = f"{get_column_letter(start_col + 1)}{row + 1}:{get_column_letter(end_col + 1)}{last_data_row + 1}"
                if not range_has_merged_cells(sheet, ref):
                    df = dataframe_from_range(value_sheet, ref)
                    structured_ranges.append(
                        TableMetadata(
                            sheet=sheet.title,
                            name=None,
                            ref=ref,
                            dataframe=df,
                            has_merged_cells=False,
                        )
                    )
                row = end_row
                continue
        row += 1
    return structured_ranges


@dataclass
class WorkbookParser:
    path: Path

    def parse(self) -> WorkbookMetadata:
        workbook = load_workbook(self.path, data_only=False, read_only=False)
        values_workbook = load_workbook(self.path, data_only=True, read_only=False)

        sheets: Dict[str, SheetMetadata] = {}
        for sheet in workbook.worksheets:
            value_sheet = values_workbook[sheet.title]
            sheet_meta = SheetMetadata(name=sheet.title)
            for row in sheet.iter_rows(values_only=False):
                for cell in row:
                    raw_formula = cell.value if cell.data_type == "f" else None
                    formula = None
                    if raw_formula:
                        formula = raw_formula if str(raw_formula).startswith("=") else f"={raw_formula}"
                    value = value_sheet[cell.coordinate].value
                    if value is None and not formula:
                        continue
                    dependencies = extract_dependencies(formula, sheet.title) if formula else []
                    cell_meta = CellMetadata(
                        sheet=sheet.title,
                        coordinate=cell.coordinate,
                        value=value,
                        formula=formula,
                        dependencies=dependencies,
                        is_input=formula is None,
                    )
                    sheet_meta.cells[cell.coordinate] = cell_meta
            sheet_meta.merged_cells = [str(range_) for range_ in sheet.merged_cells.ranges]
            for table in sheet.tables.values():
                ref = table.ref
                df = dataframe_from_range(value_sheet, ref)
                table_meta = TableMetadata(
                    sheet=sheet.title,
                    name=table.name,
                    ref=ref,
                    dataframe=df,
                    has_merged_cells=range_has_merged_cells(sheet, ref),
                )
                sheet_meta.tables.append(table_meta)
            heuristic_tables = detect_structured_ranges(sheet, value_sheet)
            existing_refs = {table.ref for table in sheet_meta.tables}
            for table_meta in heuristic_tables:
                if table_meta.ref not in existing_refs:
                    sheet_meta.tables.append(table_meta)
            sheets[sheet.title] = sheet_meta

        named_ranges = collect_named_ranges(workbook)
        return WorkbookMetadata(path=str(self.path), sheets=sheets, named_ranges=named_ranges)

    def to_json(self, metadata: WorkbookMetadata) -> str:
        def default(obj):
            if isinstance(obj, (WorkbookMetadata, SheetMetadata, CellMetadata, NamedRangeMetadata, TableMetadata)):
                data = obj.__dict__.copy()
                if isinstance(obj, TableMetadata):
                    data["dataframe"] = obj.dataframe.to_dict(orient="records")
                return data
            raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")

        return json.dumps(metadata, default=default, indent=2, ensure_ascii=False)
