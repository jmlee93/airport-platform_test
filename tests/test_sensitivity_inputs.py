import json

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from analysis.sensitivity_inputs import extract_sensitivity_inputs


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    return wb, ws


def test_detects_cells_by_value_prefix():
    wb, ws = build_workbook()
    ws["A1"] = "Sensitivity_FuelPrice"
    ws["A2"] = "NotTagged"

    records = extract_sensitivity_inputs(wb)

    assert len(records) == 1
    record = records[0]
    assert record["sheet"] == "Inputs"
    assert record["coordinate"] == "A1"
    assert record["parameter"] == "FuelPrice"
    assert record["matched_by"] == ["prefix"]


def test_detects_cells_by_comment_keyword():
    wb, ws = build_workbook()
    cell = ws["B2"]
    cell.value = 123
    cell.comment = Comment("Sensitivity_DemandGrowth: adjust assumption", "analyst")

    records = extract_sensitivity_inputs(wb)

    assert len(records) == 1
    record = records[0]
    assert record["coordinate"] == "B2"
    assert record["parameter"] == "DemandGrowth"
    assert "comment" in record["matched_by"]


def test_detects_cells_by_fill_color():
    wb, ws = build_workbook()
    cell = ws["C3"]
    cell.value = 42
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    records = extract_sensitivity_inputs(wb, highlight_colors=["FFFF00"])

    assert len(records) == 1
    record = records[0]
    assert record["coordinate"] == "C3"
    assert "fill_color" in record["matched_by"]


def test_returns_json_when_requested():
    wb, ws = build_workbook()
    ws["A1"] = "Sensitivity_Test"

    json_payload = extract_sensitivity_inputs(wb, output_format="json")
    payload = json.loads(json_payload)

    assert isinstance(json_payload, str)
    assert payload[0]["parameter"] == "Test"


@pytest.mark.parametrize("pandas_available", [True, False])
def test_dataframe_output_respects_pandas_availability(monkeypatch, pandas_available):
    wb, ws = build_workbook()
    ws["A1"] = "Sensitivity_SomeParam"

    target_attr = "analysis.sensitivity_inputs._pd"

    if pandas_available:
        class DummyDataFrame:
            def __init__(self, data):
                self.data = data

            @classmethod
            def from_records(cls, records):
                return cls(records)

        class DummyPandas:
            DataFrame = DummyDataFrame

        monkeypatch.setattr(target_attr, DummyPandas, raising=False)
        frame = extract_sensitivity_inputs(wb, output_format="dataframe")
        assert isinstance(frame, DummyDataFrame)
        assert frame.data[0]["parameter"] == "SomeParam"
    else:
        monkeypatch.setattr(target_attr, None, raising=False)
        with pytest.raises(ImportError):
            extract_sensitivity_inputs(wb, output_format="dataframe")
